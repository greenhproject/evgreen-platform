"""Construye una hoja editable para el modelo preliminar de carga DC en Panamá."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "modelo-preliminar-carga-rapida-panama.xlsx"
GREEN = "10B981"
NAVY = "0F172A"
LIGHT = "E2E8F0"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="CBD5E1")


def style_title(ws, text: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=15)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28


def style_header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_grid(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def main() -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Guía"
    style_title(guide, "Modelo preliminar — Carga rápida pública en Panamá", 6)
    guide["A3"] = "Propósito"
    guide["B3"] = "Modelo de sensibilidad para evaluar un sitio DC; no es una cotización ni una proyección de inversión."
    guide["A4"] = "Cómo usar"
    guide["B4"] = "Edite las celdas verdes de Supuestos. Las hojas de escenarios actualizan las fórmulas al abrir el archivo en Excel."
    guide["A5"] = "Advertencia"
    guide["B5"] = "Reemplazar referencias históricas de enero–agosto de 2026 con el pliego vigente y el estudio de conexión de la ubicación."
    guide["A6"] = "Fuentes"
    guide["B6"] = "ASEP tarifas y CVC 2026; propuesta de pliego 2026–2030; precio público observado B/.0,35–0,50/kWh como referencia sectorial."
    for column, width in {"A": 18, "B": 105, "C": 14, "D": 14, "E": 14, "F": 14}.items():
        guide.column_dimensions[column].width = width
    for row in range(3, 7):
        guide[f"A{row}"].font = Font(bold=True, color=NAVY)
        guide[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row].height = 34

    assumptions = wb.create_sheet("Supuestos")
    style_title(assumptions, "Supuestos editables", 5)
    headers = ["Concepto", "Valor", "Unidad", "Uso en el modelo", "Fuente / nota"]
    for index, header in enumerate(headers, 1):
        assumptions.cell(3, index, header)
    style_header(assumptions[3])
    rows = [
        ("Potencia nominal", 120, "kW", "Capacidad de cargador DC", "Supuesto; validar con diseño"),
        ("Días por mes", 30, "días", "Conversión de utilización a kWh mensuales", "Supuesto de modelación"),
        ("Eficiencia de entrega", 0.92, "%", "kWh comprados = kWh entregados / eficiencia", "Reemplazar por ficha técnica"),
        ("Demanda facturada", 120, "kW", "Cargo de demanda mensual", "Validar con simultaneidad / load management"),
        ("EDEMET energía MTD", 0.16001, "B/./kWh", "Componente energía", "ASEP ene–ago 2026"),
        ("EDEMET CVC", -0.00915, "B/./kWh", "Variación combustible", "ASEP agosto 2026; variable"),
        ("EDEMET demanda MTD", 20.62, "B/./kW-mes", "Cargo de demanda", "ASEP ene–ago 2026"),
        ("EDEMET cargo fijo", 14.27, "B/./mes", "Cargo fijo", "ASEP ene–ago 2026"),
        ("Precio público bajo", 0.35, "B/./kWh", "Escenario conservador", "Banda observada a validar"),
        ("Precio público base", 0.425, "B/./kWh", "Escenario base", "Punto medio de banda observada"),
        ("Precio público alto", 0.50, "B/./kWh", "Escenario alto uso", "Banda observada a validar"),
        ("Utilización conservadora", 0.05, "%", "Horas equivalentes", "Escenario de sensibilidad"),
        ("Utilización base", 0.15, "%", "Horas equivalentes", "Escenario de sensibilidad"),
        ("Utilización alta", 0.30, "%", "Horas equivalentes", "Escenario de sensibilidad"),
    ]
    for r, row in enumerate(rows, 4):
        for c, value in enumerate(row, 1):
            assumptions.cell(r, c, value)
    style_grid(assumptions, 4, 17, 5)
    for row in range(4, 18):
        cell = assumptions.cell(row, 2)
        cell.fill = PatternFill("solid", fgColor="D1FAE5")
        cell.font = Font(bold=True, color=NAVY)
    for row in (6, 15, 16, 17):
        assumptions.cell(row, 2).number_format = "0.0%"
    for row in range(8, 15):
        assumptions.cell(row, 2).number_format = 'B/.0.00000'
    for col, width in {"A": 29, "B": 16, "C": 14, "D": 36, "E": 42}.items():
        assumptions.column_dimensions[col].width = width
    assumptions.freeze_panes = "A4"

    scenarios = wb.create_sheet("Escenarios EDEMET")
    style_title(scenarios, "Escenarios — Estación DC 120 kW · EDEMET", 11)
    cols = [
        "Escenario", "Utilización", "Precio de venta", "kWh entregados/mes", "kWh comprados/mes",
        "Ingreso/mes", "Energía + CVC", "Demanda + fijo", "Costo total de red", "Costo efectivo/kWh", "Contribución tras red"
    ]
    for index, header in enumerate(cols, 1):
        scenarios.cell(3, index, header)
    style_header(scenarios[3])
    scenario_rows = [("Conservador", "=Supuestos!B15", "=Supuestos!B12"), ("Base", "=Supuestos!B16", "=Supuestos!B13"), ("Alto uso", "=Supuestos!B17", "=Supuestos!B14")]
    for r, (name, utilization, price) in enumerate(scenario_rows, 4):
        scenarios.cell(r, 1, name)
        scenarios.cell(r, 2, utilization)
        scenarios.cell(r, 3, price)
        scenarios.cell(r, 4, f"=Supuestos!B4*24*Supuestos!B5*B{r}")
        scenarios.cell(r, 5, f"=D{r}/Supuestos!B6")
        scenarios.cell(r, 6, f"=D{r}*C{r}")
        scenarios.cell(r, 7, f"=E{r}*(Supuestos!B8+Supuestos!B9)")
        scenarios.cell(r, 8, f"=Supuestos!B7*Supuestos!B10+Supuestos!B11")
        scenarios.cell(r, 9, f"=G{r}+H{r}")
        scenarios.cell(r, 10, f"=I{r}/D{r}")
        scenarios.cell(r, 11, f"=F{r}-I{r}")
    style_grid(scenarios, 4, 6, 11)
    for cell in scenarios["B4:B6"]:
        for item in cell:
            item.number_format = "0.0%"
    for row in scenarios.iter_rows(min_row=4, max_row=6, min_col=3, max_col=11):
        for cell in row:
            cell.number_format = 'B/.#,##0.00;[Red]-B/.#,##0.00'
    for col in range(1, 12):
        scenarios.column_dimensions[get_column_letter(col)].width = 18
    scenarios.column_dimensions["A"].width = 18
    scenarios.freeze_panes = "A4"

    zones = wb.create_sheet("Sensibilidad zonas")
    style_title(zones, "Sensibilidad territorial — caso base de 15%", 9)
    headers = ["Zona", "Energía", "CVC", "Demanda", "Fijo", "Costo efectivo/kWh", "Ingreso/mes", "Costo red/mes", "Contribución tras red"]
    for index, header in enumerate(headers, 1):
        zones.cell(3, index, header)
    style_header(zones[3])
    zone_rows = [
        ("EDEMET", 0.16001, -0.00915, 20.62, 14.27),
        ("ENSA", 0.15396, 0.00052, 12.45, 9.06),
        ("EDECHI", 0.13479, -0.01510, 24.53, 14.21),
    ]
    for r, row in enumerate(zone_rows, 4):
        for c, value in enumerate(row, 1):
            zones.cell(r, c, value)
        zones.cell(r, 6, f"=(((Supuestos!B4*24*Supuestos!B5*Supuestos!B16)/Supuestos!B6)*(B{r}+C{r})+(Supuestos!B7*D{r})+E{r})/(Supuestos!B4*24*Supuestos!B5*Supuestos!B16)")
        zones.cell(r, 7, "=Supuestos!B4*24*Supuestos!B5*Supuestos!B16*Supuestos!B13")
        zones.cell(r, 8, f"=G{r}-(G{r}-((Supuestos!B4*24*Supuestos!B5*Supuestos!B16)/Supuestos!B6)*(B{r}+C{r})-(Supuestos!B7*D{r})-E{r})")
        zones.cell(r, 9, f"=G{r}-H{r}")
    style_grid(zones, 4, 6, 9)
    for row in zones.iter_rows(min_row=4, max_row=6, min_col=2, max_col=9):
        for cell in row:
            cell.number_format = 'B/.#,##0.000;[Red]-B/.#,##0.000'
    for col in range(1, 10):
        zones.column_dimensions[get_column_letter(col)].width = 18
    zones.column_dimensions["A"].width = 16
    zones.freeze_panes = "A4"

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
