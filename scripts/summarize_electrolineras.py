from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


INPUT = Path("/home/ubuntu/upload/ELECTROLINERAS.xlsx")
OUTPUT = Path("/home/ubuntu/green-ev-platform/docs/electrolineras_investment_metrics.json")


def find_row_values(worksheet, label: str):
    for row in worksheet.iter_rows():
        if any(cell.value == label for cell in row):
            return [cell.value for cell in row]
    raise ValueError(f"Etiqueta no encontrada: {label} en {worksheet.title}")


def find_all_row_values(worksheet, label: str):
    matches = []
    for row in worksheet.iter_rows():
        if any(cell.value == label for cell in row):
            matches.append([cell.value for cell in row])
    return matches


def numbers_after_label(row_values, label: str):
    position = row_values.index(label)
    return [value for value in row_values[position + 1 :] if isinstance(value, (int, float))]


def first_number_after_label(worksheet, label: str):
    return numbers_after_label(find_row_values(worksheet, label), label)[0]


def model_errors(workbook):
    excel_error_values = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}
    errors = []
    for worksheet in workbook.worksheets:
        for cell in worksheet._cells.values():
            if cell.value in excel_error_values:
                errors.append({"sheet": worksheet.title, "cell": cell.coordinate, "error": cell.value})
    return errors


def main():
    workbook = load_workbook(INPUT, data_only=True, read_only=False)
    parameters = workbook["Parámetros"]
    cashflow = workbook["Project Cash Flow"]
    operation = workbook["Operation & Production"]

    fx = first_number_after_label(parameters, "DÓLAR")
    wacc = first_number_after_label(parameters, "Tasa de remuneracion ")
    contract_years = first_number_after_label(parameters, "Años Contrato")
    installed_power_kw = first_number_after_label(parameters, "Potencia instalada")
    price_per_kwh = first_number_after_label(parameters, "Precio venta al público")
    energy_cost_per_kwh = first_number_after_label(parameters, "Costo energía (solar+red)")

    initial_capex_cop = first_number_after_label(cashflow, "BOOT")
    irr = first_number_after_label(cashflow, "Internal Rate of Return (IRR)")
    npv_cop = first_number_after_label(cashflow, "NPV")
    sales_with_total = numbers_after_label(find_row_values(cashflow, "Sales"), "Sales")
    sales = sales_with_total[:10]
    sales_total = sales_with_total[10]
    ebitda_with_year_zero = numbers_after_label(find_row_values(workbook["Balance Sheet & Profit and Loss"], "EBITDA"), "EBITDA")
    ebitda = ebitda_with_year_zero[1:11]
    operating_station_rows = find_all_row_values(operation, "Electrolineras en FUNCIONAMIENTO")
    operating_stations = [numbers_after_label(row, "Electrolineras en FUNCIONAMIENTO")[-1] for row in operating_station_rows]
    closing_balance = numbers_after_label(find_row_values(cashflow, "F"), "F")
    cached_errors = model_errors(workbook)

    # 0th item of Sales / EBITDA is historical year 0 and is excluded from operating summary.
    sales_by_year = [value / fx for value in sales]
    ebitda_by_year = [value / fx for value in ebitda]
    metrics = {
        "as_of": "2026-08-31",
        "currency_basis": {"primary": "COP", "display_conversion": "USD", "cop_per_usd": fx},
        "core_assumptions": {
            "wacc_or_remuneration_rate": wacc,
            "contract_years": contract_years,
            "installed_power_kw_total": installed_power_kw,
            "station_count_target": 200,
            "price_cop_per_kwh": price_per_kwh,
            "energy_cost_cop_per_kwh": energy_cost_per_kwh,
        },
        "funding": {
            "boot_capex_cop": initial_capex_cop,
            "boot_capex_usd": initial_capex_cop / fx,
            "capex_per_station_cop": initial_capex_cop / 200,
            "capex_per_station_usd": initial_capex_cop / fx / 200,
        },
        "returns": {
            "project_irr": irr,
            "npv_cop": npv_cop,
            "npv_usd": npv_cop / fx,
        },
        "operations": {
            "sales_usd_by_year": sales_by_year,
            "sales_usd_total_10_years": sales_total / fx,
            "ebitda_usd_by_year": ebitda_by_year,
            "operating_stations_by_year": operating_stations,
            "closing_cash_balance_cop_by_year": closing_balance,
            "sales_year_4_usd": sales_by_year[3],
            "sales_year_10_usd": sales_by_year[9],
            "ebitda_year_4_usd": ebitda_by_year[3],
            "ebitda_year_10_usd": ebitda_by_year[9],
        },
        "audit_flags": {
            "cached_formula_error_count": len(cached_errors),
            "cached_formula_error_sample": cached_errors[:25],
            "notes": [
                "El modelo etiqueta algunas hojas como dólares, pero los flujos y el BOOT se muestran en COP; la conversión de exhibición utiliza el parámetro DÓLAR del propio archivo.",
                "Las fórmulas con errores #NUM! aparecen después del horizonte de 10 años en la sección BOOT. No se utilizan para las cifras operativas del horizonte 0-10, pero deben corregirse antes de enviar el archivo fuente a diligencia.",
                "La tasa de 12% se muestra en el archivo como tasa de remuneración / BOOT nominal; el IRR calculado en el flujo acumulado se presenta por separado y no debe confundirse con una promesa contractual de retorno.",
            ],
        },
    }
    OUTPUT.write_text(json.dumps(metrics, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(metrics, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
