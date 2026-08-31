from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


INPUT = Path("/home/ubuntu/upload/ELECTROLINERAS.xlsx")
OUTPUT_JSON = Path("/home/ubuntu/green-ev-platform/docs/electrolineras_workbook_audit.json")
OUTPUT_MD = Path("/home/ubuntu/green-ev-platform/docs/electrolineras_workbook_audit.md")


def serialise(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def get_nonempty_cells(worksheet):
    cells = []
    max_row = 0
    max_column = 0
    for cell in worksheet._cells.values():
        if cell.value is not None:
            max_row = max(max_row, cell.row)
            max_column = max(max_column, cell.column)
            cells.append({
                "address": cell.coordinate,
                "value": serialise(cell.value),
                "formula": cell.value if is_formula(cell.value) else None,
                "number_format": cell.number_format,
            })
    return cells, max_row, max_column


def preview_rows(worksheet, max_row, max_column, limit=200):
    rows = []
    for row_index in range(1, max_row + 1):
        values = [serialise(worksheet._cells.get((row_index, col_index)).value) if worksheet._cells.get((row_index, col_index)) else None for col_index in range(1, max_column + 1)]
        if any(value is not None for value in values):
            rows.append({"row": row_index, "values": values})
        if len(rows) >= limit:
            break
    return rows


def main():
    if not INPUT.exists():
        raise FileNotFoundError(f"No se encontró el archivo de entrada: {INPUT}")

    formulas = load_workbook(INPUT, data_only=False, read_only=False)
    displayed = load_workbook(INPUT, data_only=True, read_only=False)
    report = {"input": str(INPUT), "sheets": []}

    for worksheet in formulas.worksheets:
        displayed_sheet = displayed[worksheet.title]
        cells, max_row, max_column = get_nonempty_cells(worksheet)
        formula_cells = [cell for cell in cells if cell["formula"]]
        report["sheets"].append({
            "title": worksheet.title,
            "dimensions": f"A1:{get_column_letter(max_column)}{max_row}" if max_column else "Sin datos",
            "max_row": max_row,
            "max_column": max_column,
            "nonempty_cells": len(cells),
            "formula_count": len(formula_cells),
            "formulas": formula_cells,
            "preview_rows": preview_rows(displayed_sheet, max_row, max_column) if max_column else [],
        })

    OUTPUT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = ["# Auditoría estructural del libro ELECTROLINERAS", "", f"**Archivo:** `{INPUT.name}`", ""]
    for sheet in report["sheets"]:
        lines.extend([
            f"## {sheet['title']}",
            "",
            f"- Rango usado: `{sheet['dimensions']}`",
            f"- Celdas no vacías: {sheet['nonempty_cells']}",
            f"- Fórmulas: {sheet['formula_count']}",
            "",
            "| Fila | Valores visibles |",
            "|---:|---|",
        ])
        for row in sheet["preview_rows"]:
            values = " · ".join(str(value) for value in row["values"] if value is not None)
            lines.append(f"| {row['row']} | {values.replace('|', '\\|')} |")
        lines.append("")
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Auditoría escrita en {OUTPUT_JSON} y {OUTPUT_MD}")


if __name__ == "__main__":
    main()
